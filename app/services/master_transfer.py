from __future__ import annotations

import json
import secrets
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Callable

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError

from ..extensions import db
from ..models import AuditAction, Cluster, Committee, CommitteeMember, DA, PC, PM, User, Village
from .audit import model_snapshot, record_audit


TEMPLATE_VERSION = "1.0"
META_SHEET = "__META__"
STAGE_TTL = timedelta(minutes=30)


class MasterTransferError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ColumnSpec:
    header: str
    attr: str | None = None
    kind: str = "text"
    required: bool = False
    editable: bool = True
    relationship_model: type | None = None
    context: Callable[[Any], Any] | None = None


@dataclass(frozen=True, slots=True)
class ResourceSpec:
    key: str
    label: str
    model: type
    sheet: str
    columns: tuple[ColumnSpec, ...]
    move_attrs: tuple[str, ...] = ()


RESOURCES: dict[str, ResourceSpec] = {
    "pms": ResourceSpec(
        "pms", "Program Managers", PM, "PMs",
        (
            ColumnSpec("ID", "id", "integer", editable=False),
            ColumnSpec("Name", "full_name", required=True),
            ColumnSpec("Email", "email"),
            ColumnSpec("Mobile", "mobile"),
            ColumnSpec("Notes", "notes"),
            ColumnSpec("Enabled", "is_enabled", "boolean", required=True),
        ),
    ),
    "pcs": ResourceSpec(
        "pcs", "Project Coordinators", PC, "PCs",
        (
            ColumnSpec("ID", "id", "integer", editable=False),
            ColumnSpec("Name", "full_name", required=True),
            ColumnSpec("Cluster", "cluster", "cluster", required=True),
            ColumnSpec("Email", "email"),
            ColumnSpec("Mobile", "mobile"),
            ColumnSpec("Notes", "notes"),
            ColumnSpec("Enabled", "is_enabled", "boolean", required=True),
        ),
    ),
    "das": ResourceSpec(
        "das", "Development Agents", DA, "DAs",
        (
            ColumnSpec("ID", "id", "integer", editable=False),
            ColumnSpec("Name", "full_name", required=True),
            ColumnSpec("PC ID", "pc_id", "relationship", required=True, relationship_model=PC),
            ColumnSpec("PC Name", editable=False, context=lambda row: row.pc.full_name),
            ColumnSpec("Email", "email"),
            ColumnSpec("Mobile", "mobile"),
            ColumnSpec("Notes", "notes"),
            ColumnSpec("Enabled", "is_enabled", "boolean", required=True),
        ),
        move_attrs=("pc_id",),
    ),
    "villages": ResourceSpec(
        "villages", "Villages", Village, "Villages",
        (
            ColumnSpec("ID", "id", "integer", editable=False),
            ColumnSpec("Code", "code"),
            ColumnSpec("Name", "name", required=True),
            ColumnSpec("DA ID", "da_id", "relationship", required=True, relationship_model=DA),
            ColumnSpec("DA Name", editable=False, context=lambda row: row.da.full_name),
            ColumnSpec("GP Name", "gp_name"),
            ColumnSpec("District", "district"),
            ColumnSpec("Mandal", "mandal"),
            ColumnSpec("Latitude", "latitude", "number"),
            ColumnSpec("Longitude", "longitude", "number"),
            ColumnSpec("Notes", "notes"),
            ColumnSpec("Enabled", "is_enabled", "boolean", required=True),
        ),
        move_attrs=("da_id",),
    ),
    "committees": ResourceSpec(
        "committees", "Committees", Committee, "Committees",
        (
            ColumnSpec("ID", "id", "integer", editable=False),
            ColumnSpec("Village ID", "village_id", "relationship", required=True, relationship_model=Village),
            ColumnSpec("Village", editable=False, context=lambda row: row.village.name),
            ColumnSpec("Name", "name", required=True),
            ColumnSpec("Committee Type", "committee_type"),
            ColumnSpec("Notes", "notes"),
            ColumnSpec("Enabled", "is_enabled", "boolean", required=True),
        ),
        move_attrs=("village_id",),
    ),
    "committee-members": ResourceSpec(
        "committee-members", "Committee Members", CommitteeMember, "Committee Members",
        (
            ColumnSpec("ID", "id", "integer", editable=False),
            ColumnSpec("Committee ID", "committee_id", "relationship", required=True, relationship_model=Committee),
            ColumnSpec("Village", editable=False, context=lambda row: row.committee.village.name),
            ColumnSpec("Committee", editable=False, context=lambda row: row.committee.name),
            ColumnSpec("Name", "full_name", required=True),
            ColumnSpec("Gender", "gender", "gender"),
            ColumnSpec("Designation", "designation", "designation"),
            ColumnSpec("Mobile", "mobile"),
            ColumnSpec("Notes", "notes"),
            ColumnSpec("Enabled", "is_enabled", "boolean", required=True),
        ),
        move_attrs=("committee_id",),
    ),
}


def resource_catalog() -> list[dict[str, str]]:
    return [{"key": item.key, "label": item.label} for item in RESOURCES.values()]


def _resource(key: str) -> ResourceSpec:
    try:
        return RESOURCES[key]
    except KeyError as exc:
        raise MasterTransferError("Unknown master-data resource.") from exc


def _display(value: Any) -> Any:
    if hasattr(value, "value"):
        return value.value
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def _preview_value(value: Any) -> Any:
    if hasattr(value, "value"):
        return value.value
    return value


def _export_rows(spec: ResourceSpec) -> list[Any]:
    stmt = (
        db.select(spec.model)
        .where(spec.model.is_deleted.is_(False))
        .order_by(spec.model.id)
    )
    return list(db.session.scalars(stmt).unique())


def _style_export(ws, spec: ResourceSpec, row_count: int) -> None:
    navy, white, system, editable = "1F4E5F", "FFFFFF", "E8E7E1", "FFFFFF"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.protection = Protection(locked=True)
    for index, col in enumerate(spec.columns, 1):
        for row in range(2, row_count + 2):
            cell = ws.cell(row, index)
            if col.editable:
                cell.fill = PatternFill("solid", fgColor=editable)
                cell.protection = Protection(locked=False)
            else:
                cell.fill = PatternFill("solid", fgColor=system)
                cell.font = Font(color="555555")
                cell.protection = Protection(locked=True)
        sample = [str(ws.cell(row, index).value or "") for row in range(1, min(row_count + 1, 150) + 1)]
        ws.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, sample), default=10) + 2, 11), 34)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    ws.protection.enable()


def build_export_workbook(resource_key: str, user: User) -> Workbook:
    spec = _resource(resource_key)
    rows = _export_rows(spec)
    wb = Workbook()
    ws = wb.active
    ws.title = spec.sheet
    ws.append([col.header for col in spec.columns])
    for record in rows:
        values = []
        for col in spec.columns:
            if col.context:
                value = col.context(record)
            elif col.attr:
                value = getattr(record, col.attr)
            else:
                value = None
            values.append(_display(value))
        ws.append(values)
        for cell in ws[ws.max_row]:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.data_type = "s"
    # Add one unlocked blank row so new records can be added without copying formatting.
    ws.append([None for _ in spec.columns])
    _style_export(ws, spec, len(rows) + 1)

    meta = wb.create_sheet(META_SHEET)
    meta.sheet_state = "hidden"
    for key, value in (
        ("template_version", TEMPLATE_VERSION),
        ("resource", spec.key),
        ("exported_for_user_id", user.id),
        ("exported_at_utc", datetime.now(timezone.utc).isoformat()),
    ):
        meta.append([key, value])
    return wb


def _text(value: Any) -> str | None:
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _boolean(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    raw = str(value).strip().casefold()
    if raw in {"yes", "true", "1", "enabled", "y"}:
        return True
    if raw in {"no", "false", "0", "disabled", "n"}:
        return False
    raise MasterTransferError("Enabled must be Yes or No.")


def _convert(col: ColumnSpec, value: Any) -> Any:
    if value in (None, ""):
        return None
    if col.kind == "text":
        return _text(value)
    if col.kind in {"integer", "relationship"}:
        try:
            return int(value)
        except (TypeError, ValueError) as exc:
            raise MasterTransferError(f"{col.header} must be an integer ID.") from exc
    if col.kind == "number":
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise MasterTransferError(f"{col.header} must be a number.") from exc
    if col.kind == "boolean":
        return _boolean(value)
    if col.kind == "cluster":
        raw = _text(value)
        try:
            return Cluster(raw) if raw else None
        except ValueError as exc:
            raise MasterTransferError("Cluster must be CSRB or PDTC.") from exc
    if col.kind == "designation":
        raw = _text(value)
        if raw and raw not in {"President", "Vice President", "Secretary", "Member"}:
            raise MasterTransferError("Designation must be President, Vice President, Secretary or Member.")
        return raw
    if col.kind == "gender":
        raw = _text(value)
        if raw and raw not in {"Male", "Female", "Unknown", "Other", "Not stated"}:
            raise MasterTransferError("Gender must be Male, Female, Unknown, Other or Not stated.")
        return raw
    return value


def _headers(ws, spec: ResourceSpec) -> dict[str, int]:
    found = {_text(cell.value): index for index, cell in enumerate(ws[1], 1) if _text(cell.value)}
    missing = [col.header for col in spec.columns if col.header not in found]
    if missing:
        raise MasterTransferError("Workbook is missing columns: " + ", ".join(missing))
    return found


def _meta(wb, spec: ResourceSpec, user: User) -> None:
    if META_SHEET not in wb.sheetnames:
        raise MasterTransferError("This is not a Model Village master-data export workbook.")
    values = {
        _text(row[0].value): _text(row[1].value)
        for row in wb[META_SHEET].iter_rows(min_row=1, max_col=2)
        if _text(row[0].value)
    }
    if values.get("template_version") != TEMPLATE_VERSION:
        raise MasterTransferError("Unsupported master-data template version.")
    if values.get("resource") != spec.key:
        raise MasterTransferError("The selected resource does not match this workbook.")
    if values.get("exported_for_user_id") != str(user.id):
        raise MasterTransferError("This workbook was exported for a different login.")


def _validate_relationship(col: ColumnSpec, value: Any) -> str | None:
    if value is None or not col.relationship_model:
        return None
    related = db.session.get(col.relationship_model, value)
    if related is None or getattr(related, "is_deleted", False):
        return f"{col.header} does not reference an available record."
    return None


def _duplicate_error(spec: ResourceSpec, values: dict[str, Any], record_id: int | None) -> str | None:
    model = spec.model
    filters = []
    if model in {PM, PC, DA}:
        for attr in ("email", "mobile"):
            value = values.get(attr)
            if value:
                row = db.session.scalar(db.select(model).where(getattr(model, attr) == value))
                if row and row.id != record_id:
                    return f"{attr.title()} is already used by another record."
    elif model is Village:
        if values.get("code"):
            row = db.session.scalar(db.select(Village).where(Village.code == values["code"]))
            if row and row.id != record_id:
                return "Village Code is already used by another record."
        if values.get("da_id") and values.get("name"):
            row = db.session.scalar(
                db.select(Village).where(Village.da_id == values["da_id"], Village.name == values["name"])
            )
            if row and row.id != record_id:
                return "That DA already has a Village with this Name."
    elif model is Committee and values.get("village_id") and values.get("name"):
        row = db.session.scalar(
            db.select(Committee).where(
                Committee.village_id == values["village_id"],
                Committee.name == values["name"],
            )
        )
        if row and row.id != record_id:
            return "That Village already has a Committee with this Name."
    return None


def _workbook_signatures(spec: ResourceSpec, values: dict[str, Any]) -> list[tuple[str, tuple[Any, ...]]]:
    signatures: list[tuple[str, tuple[Any, ...]]] = []
    if spec.model in {PM, PC, DA}:
        for attr in ("email", "mobile"):
            value = values.get(attr)
            if value:
                signatures.append((attr.title(), (attr, str(value).casefold())))
    elif spec.model is Village:
        if values.get("code"):
            signatures.append(("Village Code", ("code", str(values["code"]).casefold())))
        if values.get("da_id") and values.get("name"):
            signatures.append(
                ("DA + Village Name", ("da_name", int(values["da_id"]), str(values["name"]).casefold()))
            )
    elif spec.model is Committee and values.get("village_id") and values.get("name"):
        signatures.append(
            ("Village + Committee Name", ("village_committee", int(values["village_id"]), str(values["name"]).casefold()))
        )
    return signatures


def preview_import(path: Path, resource_key: str, user: User) -> dict[str, Any]:
    spec = _resource(resource_key)
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        raise MasterTransferError("The uploaded file is not a readable .xlsx workbook.") from exc
    _meta(wb, spec, user)
    if spec.sheet not in wb.sheetnames:
        raise MasterTransferError(f"Workbook must contain the '{spec.sheet}' sheet.")
    ws = wb[spec.sheet]
    headers = _headers(ws, spec)
    parsed = []
    seen_ids: set[int] = set()

    for excel_row in range(2, ws.max_row + 1):
        raw_values = {col.header: ws.cell(excel_row, headers[col.header]).value for col in spec.columns}
        if not any(value not in (None, "") for value in raw_values.values()):
            continue

        errors: list[str] = []
        record_id = None
        raw_id = raw_values.get("ID")
        if raw_id not in (None, ""):
            try:
                record_id = int(raw_id)
            except (TypeError, ValueError):
                errors.append("ID is invalid.")
            if record_id in seen_ids:
                errors.append("ID appears more than once in this workbook.")
            if record_id:
                seen_ids.add(record_id)

        record = db.session.get(spec.model, record_id) if record_id else None
        if record_id and (record is None or getattr(record, "is_deleted", False)):
            errors.append("ID does not reference an available record.")

        values: dict[str, Any] = {}
        for col in spec.columns:
            if not col.editable or not col.attr:
                continue
            try:
                converted = _convert(col, raw_values[col.header])
            except MasterTransferError as exc:
                converted = None
                errors.append(str(exc))
            if col.required and converted is None:
                errors.append(f"{col.header} is required.")
            rel_error = _validate_relationship(col, converted)
            if rel_error:
                errors.append(rel_error)
            values[col.attr] = converted

        dup = _duplicate_error(spec, values, record_id)
        if dup:
            errors.append(dup)

        if errors:
            action = "Error"
        elif record is None:
            action = "New"
        else:
            changed_attrs = [
                attr for attr, value in values.items()
                if getattr(record, attr) != value
            ]
            if not changed_attrs:
                action = "Unchanged"
            elif any(attr in spec.move_attrs for attr in changed_attrs):
                action = "Moved"
            else:
                action = "Changed"

        parsed.append(
            {
                "excel_row": excel_row,
                "id": record_id,
                "name": values.get("full_name") or values.get("name") or (
                    getattr(record, "full_name", None) if record else None
                ) or (getattr(record, "name", None) if record else None) or "New record",
                "action": action,
                "errors": errors,
                "values": {
                    key: (_preview_value(value) if value is not None else None)
                    for key, value in values.items()
                },
            }
        )

    # Database checks above cannot detect two brand-new rows that conflict with each other.
    # Flag both rows before the user is allowed to confirm.
    seen_signatures: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in parsed:
        for label, signature in _workbook_signatures(spec, row["values"]):
            previous = seen_signatures.get(signature)
            if previous and previous.get("id") != row.get("id"):
                message = f"{label} is duplicated inside this workbook."
                if message not in row["errors"]:
                    row["errors"].append(message)
                if message not in previous["errors"]:
                    previous["errors"].append(message)
                row["action"] = "Error"
                previous["action"] = "Error"
            else:
                seen_signatures[signature] = row

    counts = {key: 0 for key in ("New", "Changed", "Moved", "Unchanged", "Error")}
    for row in parsed:
        counts[row["action"]] += 1
    return {
        "resource": spec.key,
        "label": spec.label,
        "counts": counts,
        "has_errors": counts["Error"] > 0,
        "rows": parsed,
    }


def _stage_root() -> Path:
    root = Path(current_app.instance_path) / "import_staging" / "master-data"
    root.mkdir(parents=True, exist_ok=True)
    return root


def stage_import(path: Path, resource_key: str, user: User) -> tuple[str, dict[str, Any]]:
    preview = preview_import(path, resource_key, user)
    token = secrets.token_urlsafe(24)
    root = _stage_root()
    (root / f"{token}.xlsx").write_bytes(path.read_bytes())
    (root / f"{token}.json").write_text(
        json.dumps(
            {
                "user_id": user.id,
                "resource": resource_key,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
        ),
        encoding="utf-8",
    )
    return token, preview


def _staged(token: str, resource_key: str, user: User) -> Path:
    if not token or any(ch not in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_" for ch in token):
        raise MasterTransferError("Invalid import preview token.")
    root = _stage_root()
    meta_path, data_path = root / f"{token}.json", root / f"{token}.xlsx"
    if not meta_path.exists() or not data_path.exists():
        raise MasterTransferError("Import preview expired or was not found.")
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        created = datetime.fromisoformat(meta["created_at"])
    except Exception as exc:
        raise MasterTransferError("Import preview metadata is invalid.") from exc
    now = datetime.now(timezone.utc)
    if created.tzinfo is None:
        created = created.replace(tzinfo=timezone.utc)
    if now - created > STAGE_TTL:
        meta_path.unlink(missing_ok=True)
        data_path.unlink(missing_ok=True)
        raise MasterTransferError("Import preview expired. Validate the workbook again.")
    if int(meta.get("user_id", -1)) != user.id or meta.get("resource") != resource_key:
        raise MasterTransferError("Import preview does not belong to this user/resource.")
    return data_path


def _apply_values(record: Any, values: dict[str, Any], spec: ResourceSpec) -> None:
    for col in spec.columns:
        if not col.editable or not col.attr or col.attr not in values:
            continue
        value = values[col.attr]
        if col.kind == "cluster" and value is not None:
            value = Cluster(value)
        elif col.kind == "boolean" and value is not None and not isinstance(value, bool):
            value = _boolean(value)
        setattr(record, col.attr, value)


def confirm_import(token: str, resource_key: str, user: User) -> dict[str, Any]:
    spec = _resource(resource_key)
    path = _staged(token, resource_key, user)
    preview = preview_import(path, resource_key, user)
    if preview["has_errors"]:
        raise MasterTransferError("Import still contains validation errors; nothing was saved.")

    created = updated = moved = 0
    try:
        for row in preview["rows"]:
            if row["action"] not in {"New", "Changed", "Moved"}:
                continue
            values = row["values"]
            record = db.session.get(spec.model, row["id"]) if row["id"] else None
            if record is None:
                record = spec.model()
                _apply_values(record, values, spec)
                db.session.add(record)
                db.session.flush()
                record_audit(AuditAction.IMPORT, record, after=model_snapshot(record), actor=user)
                created += 1
            else:
                before = model_snapshot(record)
                _apply_values(record, values, spec)
                db.session.flush()
                action = AuditAction.MOVE if row["action"] == "Moved" else AuditAction.IMPORT
                record_audit(action, record, before=before, after=model_snapshot(record), actor=user)
                if row["action"] == "Moved":
                    moved += 1
                else:
                    updated += 1
        db.session.commit()
    except IntegrityError as exc:
        db.session.rollback()
        raise MasterTransferError(
            "The database rejected one or more changes because they conflict with existing master data. "
            "Validate a fresh export and try again."
        ) from exc
    except Exception:
        db.session.rollback()
        raise
    finally:
        root = _stage_root()
        for suffix in (".xlsx", ".json"):
            (root / f"{token}{suffix}").unlink(missing_ok=True)

    return {
        "resource": resource_key,
        "created": created,
        "updated": updated,
        "moved": moved,
    }
