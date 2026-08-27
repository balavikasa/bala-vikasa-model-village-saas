from __future__ import annotations

import json
import secrets
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from ..extensions import db
from ..models import (
    DA,
    ActionPlan,
    ActionPlanType,
    AuditAction,
    Committee,
    User,
    Village,
)
from ..scoping import can_manage_action_plan, require_scoped, scoped_select
from ..timeutils import current_month, local_today
from .audit import model_snapshot, record_audit

TEMPLATE_VERSION = "2.0"
STAGE_TTL_MINUTES = 30
EXPORT_SHEET = "Action Plans"
META_SHEET = "__META__"
PLAN_TYPES = tuple(member.value for member in ActionPlanType)
SYSTEM_HEADERS = ("Plan ID", "Month", "Committee ID", "DA", "Village", "Committee", "Status")
EDITABLE_HEADERS = ("Type", "Assigned Date", "Notes")
ALL_HEADERS = (
    "Plan ID",
    "Month",
    "Committee ID",
    "DA",
    "Village",
    "Committee",
    "Type",
    "Assigned Date",
    "Notes",
    "Status",
)


class PlanningError(ValueError):
    pass


@dataclass(slots=True)
class ImportRow:
    excel_row: int
    committee_id: int | None
    plan_id: int | None
    plan_type: ActionPlanType | None
    assigned_date: date | None
    notes: str | None
    action: str
    errors: list[str]
    da_name: str = ""
    village_name: str = ""
    committee_name: str = ""


def month_start(value: str | date | None = None) -> date:
    if isinstance(value, date):
        return value.replace(day=1)
    raw = str(value or "").strip()
    if not raw:
        return current_month()
    try:
        if len(raw) == 7:
            return date.fromisoformat(f"{raw}-01")
        return date.fromisoformat(raw).replace(day=1)
    except ValueError as exc:
        raise PlanningError("Month must use YYYY-MM format.") from exc


def month_key(value: date) -> str:
    return value.strftime("%Y-%m")


def month_label(value: date) -> str:
    return value.strftime("%B %Y")


def add_months(value: date, months: int) -> date:
    index = value.year * 12 + (value.month - 1) + months
    return date(index // 12, index % 12 + 1, 1)


def action_plan_status(plan: ActionPlan, today: date | None = None) -> str:
    today = today or local_today()
    if not plan.is_executable:
        return "Draft"
    if plan.attendance_entry and not plan.attendance_entry.is_deleted:
        return plan.attendance_entry.status.value
    if plan.specials_entry and not plan.specials_entry.is_deleted:
        return plan.specials_entry.status.value if plan.specials_entry.status else "Completed"
    if plan.assigned_date and plan.assigned_date < today:
        return "Failure"
    if plan.assigned_date == today:
        return "Due today"
    return "Scheduled"


def plan_locked(plan: ActionPlan, today: date | None = None) -> bool:
    """Monthly history is immutable once completed or its month has passed."""
    today = today or local_today()
    current = today.replace(day=1)
    if plan.attendance_entry and not plan.attendance_entry.is_deleted:
        return True
    if plan.specials_entry and not plan.specials_entry.is_deleted:
        return True
    if plan.assigned_date and plan.assigned_date < today:
        return True
    return bool(plan.plan_month and plan.plan_month < current)


def _committee_query(user: User):
    return (
        scoped_select(Committee, user)
        .options(
            selectinload(Committee.village)
            .selectinload(Village.da)
            .selectinload(DA.pc)
        )
        .order_by(Committee.name)
    )


def visible_committees(user: User) -> list[Committee]:
    rows = list(db.session.scalars(_committee_query(user)).unique())
    return sorted(rows, key=lambda committee: (
        committee.village.da.full_name.casefold(),
        committee.village.name.casefold(),
        committee.name.casefold(),
    ))


def monthly_plans(user: User, month: date) -> dict[int, ActionPlan]:
    stmt = (
        scoped_select(ActionPlan, user)
        .where(ActionPlan.plan_month == month)
        .options(
            selectinload(ActionPlan.committee)
            .selectinload(Committee.village)
            .selectinload(Village.da)
            .selectinload(DA.pc),
            selectinload(ActionPlan.attendance_entry),
            selectinload(ActionPlan.specials_entry),
        )
    )
    return {plan.committee_id: plan for plan in db.session.scalars(stmt).unique()}


def _monthly_plan_slots(user: User, month: date) -> dict[int, ActionPlan]:
    """Return every scoped committee/month row, including archived lifecycle rows.

    The database uniqueness constraint covers soft-deleted ActionPlan rows too.
    Import and month preparation therefore need a lifecycle-inclusive lookup so
    they never attempt a second physical row for the same committee/month.
    """
    stmt = (
        scoped_select(
            ActionPlan,
            user,
            include_deleted=True,
            include_disabled=True,
        )
        .where(ActionPlan.plan_month == month)
        .options(
            selectinload(ActionPlan.committee)
            .selectinload(Committee.village)
            .selectinload(Village.da)
            .selectinload(DA.pc),
            selectinload(ActionPlan.attendance_entry),
            selectinload(ActionPlan.specials_entry),
        )
    )
    return {plan.committee_id: plan for plan in db.session.scalars(stmt).unique()}


def _plan_has_any_field_history(plan: ActionPlan) -> bool:
    """Return True even when a linked field row is itself soft-deleted."""
    return plan.attendance_entry is not None or plan.specials_entry is not None


def planning_rows(user: User, month: date) -> list[dict[str, Any]]:
    plans = monthly_plans(user, month)
    rows: list[dict[str, Any]] = []
    for committee in visible_committees(user):
        plan = plans.get(committee.id)
        village = committee.village
        da = village.da
        rows.append(
            {
                "plan_id": plan.id if plan else None,
                "month": month_key(month),
                "committee_id": committee.id,
                "da_id": da.id,
                "da_name": da.full_name,
                "village_id": village.id,
                "village_name": village.name,
                "committee_name": committee.name,
                "plan_type": plan.plan_type.value if plan and plan.plan_type else None,
                "assigned_date": plan.assigned_date.isoformat() if plan and plan.assigned_date else None,
                "notes": plan.notes if plan else None,
                "status": action_plan_status(plan) if plan else "Draft",
                "locked": plan_locked(plan) if plan else False,
                "entry_id": (
                    plan.attendance_entry.id
                    if plan and plan.attendance_entry and not plan.attendance_entry.is_deleted
                    else plan.specials_entry.id
                    if plan and plan.specials_entry and not plan.specials_entry.is_deleted
                    else None
                ),
            }
        )
    return rows


def planning_summary(rows: Iterable[dict[str, Any]]) -> dict[str, int]:
    result = {
        "total": 0,
        "draft": 0,
        "scheduled": 0,
        "due_today": 0,
        "early": 0,
        "on_time": 0,
        "postponed": 0,
        "failure": 0,
    }
    for row in rows:
        result["total"] += 1
        key = {
            "Draft": "draft",
            "Scheduled": "scheduled",
            "Due today": "due_today",
            "Early": "early",
            "On-time": "on_time",
            "Postponed": "postponed",
            "Failure": "failure",
        }.get(row["status"])
        if key:
            result[key] += 1
    return result


def _cell_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_type(value: Any) -> ActionPlanType | None:
    raw = _cell_text(value)
    if raw is None:
        return None
    for member in ActionPlanType:
        if raw.casefold() in {member.value.casefold(), member.name.casefold()}:
            return member
    raise PlanningError("Type must be Attendance or Specials.")


def _parse_excel_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    for candidate in (raw, raw.replace("/", "-")):
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y"):
            try:
                return datetime.strptime(candidate, fmt).date()
            except ValueError:
                pass
    raise PlanningError("Assigned Date must be a valid date.")


def _headers(ws) -> dict[str, int]:
    result = {}
    for index, cell in enumerate(ws[1], 1):
        label = _cell_text(cell.value)
        if label:
            result[label] = index
    missing = [header for header in ALL_HEADERS if header not in result]
    if missing:
        raise PlanningError("Workbook is missing columns: " + ", ".join(missing))
    return result


def _export_styles(ws, row_count: int) -> None:
    navy = "1F4E5F"
    white = "FFFFFF"
    system = "E8E7E1"
    required = "FFF2CC"
    editable = "FFFFFF"
    locked_font = "555555"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.protection = Protection(locked=True)
    headers = {cell.value: cell.column for cell in ws[1]}
    for name in SYSTEM_HEADERS:
        col = headers[name]
        for row in range(2, row_count + 2):
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor=system)
            c.font = Font(color=locked_font)
            c.protection = Protection(locked=True)
    for name in ("Type", "Assigned Date"):
        col = headers[name]
        for row in range(2, row_count + 2):
            c = ws.cell(row, col)
            c.fill = PatternFill("solid", fgColor=required)
            c.protection = Protection(locked=False)
    notes_col = headers["Notes"]
    for row in range(2, row_count + 2):
        ws.cell(row, notes_col).fill = PatternFill("solid", fgColor=editable)
        ws.cell(row, notes_col).protection = Protection(locked=False)

    type_col = get_column_letter(headers["Type"])
    validation = DataValidation(
        type="list",
        formula1='"Attendance,Specials"',
        allow_blank=True,
        error="Choose Attendance or Specials.",
        errorTitle="Invalid action-plan type",
    )
    ws.add_data_validation(validation)
    validation.add(f"{type_col}2:{type_col}{row_count + 1}")

    assigned_col = get_column_letter(headers["Assigned Date"])
    type_col_letter = get_column_letter(headers["Type"])
    yellow = PatternFill("solid", fgColor="FFF2CC")
    ws.conditional_formatting.add(
        f"{assigned_col}2:{assigned_col}{row_count + 1}",
        FormulaRule(
            formula=[f'AND(${type_col_letter}2<>"",${assigned_col}2="")'],
            fill=yellow,
        ),
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True
    ws.protection.enable()
    # No password: protection is guidance, server-side validation is the security boundary.
    widths = {
        "Plan ID": 13,
        "Month": 13,
        "Committee ID": 14,
        "DA": 22,
        "Village": 24,
        "Committee": 30,
        "Type": 16,
        "Assigned Date": 18,
        "Notes": 38,
        "Status": 16,
    }
    for name, width in widths.items():
        ws.column_dimensions[get_column_letter(headers[name])].width = width


def build_export_workbook(user: User, month: date) -> Workbook:
    rows = planning_rows(user, month)
    wb = Workbook()
    ws = wb.active
    ws.title = EXPORT_SHEET
    ws.append(ALL_HEADERS)
    for row in rows:
        ws.append(
            [
                row["plan_id"],
                row["month"],
                row["committee_id"],
                row["da_name"],
                row["village_name"],
                row["committee_name"],
                row["plan_type"],
                row["assigned_date"],
                row["notes"],
                row["status"],
            ]
        )
    _export_styles(ws, len(rows))
    # Immutable rows are grey/locked even in the editable columns.
    headers = {cell.value: cell.column for cell in ws[1]}
    for excel_row, row in enumerate(rows, 2):
        if not row["locked"]:
            continue
        for name in EDITABLE_HEADERS:
            cell = ws.cell(excel_row, headers[name])
            cell.fill = PatternFill("solid", fgColor="E8E7E1")
            cell.font = Font(color="555555")
            cell.protection = Protection(locked=True)

    meta = wb.create_sheet(META_SHEET)
    meta.sheet_state = "hidden"
    for key, value in (
        ("template_version", TEMPLATE_VERSION),
        ("month", month_key(month)),
        ("exported_for_user_id", user.id),
        ("exported_for_role", user.role.value),
        ("exported_at_utc", datetime.now(UTC).isoformat()),
    ):
        meta.append([key, value])
    return wb


def _validate_meta(wb, user: User, selected_month: date) -> None:
    if META_SHEET not in wb.sheetnames:
        raise PlanningError("This is not a Model Village action-plan export workbook.")
    meta = {
        _cell_text(row[0].value): _cell_text(row[1].value)
        for row in wb[META_SHEET].iter_rows(min_row=1, max_col=2)
        if _cell_text(row[0].value)
    }
    if meta.get("template_version") != TEMPLATE_VERSION:
        raise PlanningError("Unsupported action-plan template version.")
    if meta.get("month") != month_key(selected_month):
        raise PlanningError(
            f"Workbook month is {meta.get('month') or 'unknown'}, not {month_key(selected_month)}."
        )
    if meta.get("exported_for_user_id") != str(user.id):
        raise PlanningError(
            "This workbook was exported for a different login. Export a fresh sheet from your own Action Plans page."
        )


def preview_import(path: Path, user: User, selected_month: date) -> dict[str, Any]:
    if selected_month < current_month():
        raise PlanningError("Past monthly history is immutable and cannot be imported.")
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as exc:
        raise PlanningError("The uploaded file is not a readable .xlsx workbook.") from exc
    _validate_meta(wb, user, selected_month)
    if EXPORT_SHEET not in wb.sheetnames:
        raise PlanningError(f"Workbook must contain a '{EXPORT_SHEET}' sheet.")

    ws = wb[EXPORT_SHEET]
    headers = _headers(ws)
    visible = {committee.id: committee for committee in visible_committees(user)}
    existing = monthly_plans(user, selected_month)
    slots = _monthly_plan_slots(user, selected_month)
    seen: set[int] = set()
    parsed: list[ImportRow] = []

    for excel_row in range(2, ws.max_row + 1):
        raw_committee = ws.cell(excel_row, headers["Committee ID"]).value
        if raw_committee in (None, ""):
            # Allow blank trailing / note rows only if the row contains no editable data.
            if not any(
                ws.cell(excel_row, headers[name]).value not in (None, "")
                for name in EDITABLE_HEADERS
            ):
                continue
        errors: list[str] = []
        committee_id = None
        try:
            committee_id = int(raw_committee)
        except (TypeError, ValueError):
            errors.append("Committee ID is missing or invalid.")

        committee = visible.get(committee_id) if committee_id else None
        if committee_id and not committee:
            errors.append("Committee is outside your permitted scope or does not exist.")
        if committee_id in seen:
            errors.append("Committee appears more than once in this workbook.")
        if committee_id:
            seen.add(committee_id)

        workbook_month = _cell_text(ws.cell(excel_row, headers["Month"]).value)
        if workbook_month != month_key(selected_month):
            errors.append(f"Month must remain {month_key(selected_month)}.")

        try:
            plan_type = _parse_type(ws.cell(excel_row, headers["Type"]).value)
        except PlanningError as exc:
            plan_type = None
            errors.append(str(exc))
        try:
            assigned = _parse_excel_date(ws.cell(excel_row, headers["Assigned Date"]).value)
        except PlanningError as exc:
            assigned = None
            errors.append(str(exc))

        if assigned is not None and plan_type is None:
            errors.append("Choose Attendance or Specials before assigning a date.")
        if assigned and (assigned.year != selected_month.year or assigned.month != selected_month.month):
            errors.append(f"Assigned Date must be inside {month_label(selected_month)}.")

        notes = _cell_text(ws.cell(excel_row, headers["Notes"]).value)
        plan = existing.get(committee_id) if committee_id else None
        archived_slot = slots.get(committee_id) if committee_id and plan is None else None

        if archived_slot is not None:
            if _plan_has_any_field_history(archived_slot):
                errors.append(
                    "An archived monthly plan with field history already exists. "
                    "Restore it from Recycle Bin instead of importing a replacement."
                )
            elif plan_locked(archived_slot):
                errors.append(
                    "An archived monthly plan occupies this committee/month and is locked. "
                    "Restore or resolve it before importing."
                )

        plan_id_raw = ws.cell(excel_row, headers["Plan ID"]).value
        try:
            workbook_plan_id = int(plan_id_raw) if plan_id_raw not in (None, "") else None
        except (TypeError, ValueError):
            workbook_plan_id = None
            errors.append("Plan ID was modified and is invalid.")

        if plan and workbook_plan_id not in (None, plan.id):
            errors.append("Plan ID does not match the database record.")
        if not plan and workbook_plan_id is not None:
            errors.append("Plan ID refers to a plan that does not exist for this month.")
        if plan and plan_locked(plan):
            changed = (
                plan.plan_type != plan_type
                or plan.assigned_date != assigned
                or (plan.notes or None) != notes
            )
            if changed:
                errors.append("This monthly plan is locked by immutable history and cannot be changed.")

        if errors:
            action = "Error"
        elif plan is None:
            action = "New" if (plan_type or assigned or notes) else "Unchanged"
        else:
            changed = (
                plan.plan_type != plan_type
                or plan.assigned_date != assigned
                or (plan.notes or None) != notes
            )
            action = "Changed" if changed else "Unchanged"

        parsed.append(
            ImportRow(
                excel_row=excel_row,
                committee_id=committee_id,
                plan_id=plan.id if plan else None,
                plan_type=plan_type,
                assigned_date=assigned,
                notes=notes,
                action=action,
                errors=errors,
                da_name=committee.village.da.full_name if committee else "",
                village_name=committee.village.name if committee else "",
                committee_name=committee.name if committee else "",
            )
        )

    counts = {key: 0 for key in ("New", "Changed", "Unchanged", "Error")}
    for row in parsed:
        counts[row.action] += 1
    return {
        "month": month_key(selected_month),
        "month_label": month_label(selected_month),
        "counts": counts,
        "has_errors": counts["Error"] > 0,
        "rows": [
            {
                "excel_row": row.excel_row,
                "committee_id": row.committee_id,
                "plan_id": row.plan_id,
                "da_name": row.da_name,
                "village_name": row.village_name,
                "committee_name": row.committee_name,
                "plan_type": row.plan_type.value if row.plan_type else None,
                "assigned_date": row.assigned_date.isoformat() if row.assigned_date else None,
                "notes": row.notes,
                "action": row.action,
                "errors": row.errors,
            }
            for row in parsed
        ],
    }


def _stage_root() -> Path:
    root = Path(current_app.instance_path) / "import_staging" / "action-plans"
    root.mkdir(parents=True, exist_ok=True)
    return root


def stage_import(source_path: Path, user: User, selected_month: date) -> tuple[str, dict[str, Any]]:
    preview = preview_import(source_path, user, selected_month)
    token = secrets.token_urlsafe(24)
    root = _stage_root()
    target = root / f"{token}.xlsx"
    target.write_bytes(source_path.read_bytes())
    (root / f"{token}.json").write_text(
        json.dumps(
            {
                "user_id": user.id,
                "month": month_key(selected_month),
                "created_at": datetime.now(UTC).isoformat(),
            }
        ),
        encoding="utf-8",
    )
    return token, preview


def _staged_file(token: str, user: User, selected_month: date) -> Path:
    if not token or any(ch not in "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-_" for ch in token):
        raise PlanningError("Invalid import preview token.")
    root = _stage_root()
    meta_path = root / f"{token}.json"
    xlsx_path = root / f"{token}.xlsx"
    if not meta_path.exists() or not xlsx_path.exists():
        raise PlanningError("Import preview expired or was not found.")
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise PlanningError("Import preview metadata is invalid.") from exc
    if int(meta.get("user_id", -1)) != user.id or meta.get("month") != month_key(selected_month):
        raise PlanningError("Import preview does not belong to this user/month.")
    try:
        created = datetime.fromisoformat(meta["created_at"])
        if created.tzinfo is None:
            created = created.replace(tzinfo=UTC)
    except Exception as exc:
        raise PlanningError("Import preview metadata is invalid.") from exc
    if datetime.now(UTC) - created > timedelta(minutes=STAGE_TTL_MINUTES):
        meta_path.unlink(missing_ok=True)
        xlsx_path.unlink(missing_ok=True)
        raise PlanningError("Import preview expired. Validate the workbook again.")
    return xlsx_path


def confirm_import(token: str, user: User, selected_month: date) -> dict[str, Any]:
    path = _staged_file(token, user, selected_month)
    preview = preview_import(path, user, selected_month)
    if preview["has_errors"]:
        raise PlanningError("Import still contains validation errors; nothing was saved.")

    # Include archived rows because the database uniqueness constraint applies to
    # them as well. A safe archived slot is restored in place so its ActionPlan ID
    # remains stable.
    existing = _monthly_plan_slots(user, selected_month)
    created = updated = 0
    committed = False

    try:
        for row in preview["rows"]:
            if row["action"] not in {"New", "Changed"}:
                continue

            committee = require_scoped(Committee, int(row["committee_id"]), user)
            if not can_manage_action_plan(user, committee):
                raise PlanningError("A row moved outside your action-plan management scope.")

            plan = existing.get(committee.id)
            plan_type = ActionPlanType(row["plan_type"]) if row["plan_type"] else None
            assigned = date.fromisoformat(row["assigned_date"]) if row["assigned_date"] else None

            if plan is None:
                plan = ActionPlan(
                    committee_id=committee.id,
                    title=committee.name,
                    description=None,
                    plan_month=selected_month,
                    plan_type=plan_type,
                    assigned_date=assigned,
                    assigned_by_user_id=user.id,
                    notes=row["notes"],
                )
                db.session.add(plan)
                try:
                    db.session.flush()
                except IntegrityError as exc:
                    raise PlanningError(
                        "Another action plan already occupies this committee/month. "
                        "Validate the workbook again before confirming."
                    ) from exc

                record_audit(
                    AuditAction.IMPORT,
                    plan,
                    after=model_snapshot(plan),
                    actor=user,
                )
                existing[committee.id] = plan
                created += 1
                continue

            was_archived = bool(plan.is_deleted or not plan.is_enabled)
            if was_archived:
                # Never revive a plan that has any field-entry history, even if the
                # field row is also in Recycle Bin. Doing so would make Reports say
                # "no active entry" while the DB unique key still blocks a new DA
                # submission.
                if _plan_has_any_field_history(plan):
                    raise PlanningError(
                        "An archived action plan with field history cannot be reused by import. "
                        "Restore it from Recycle Bin instead."
                    )
                if plan_locked(plan):
                    raise PlanningError(
                        "An archived action plan for this committee/month is locked and cannot be reused."
                    )
            elif plan_locked(plan):
                raise PlanningError(
                    "A monthly plan became locked before import confirmation."
                )

            before = model_snapshot(plan)

            if was_archived:
                plan.is_deleted = False
                plan.deleted_at = None
                plan.is_enabled = True

            plan.title = committee.name
            plan.plan_type = plan_type
            plan.assigned_date = assigned
            plan.assigned_by_user_id = user.id
            plan.notes = row["notes"]

            record_audit(
                AuditAction.IMPORT,
                plan,
                before=before,
                after=model_snapshot(plan),
                actor=user,
            )

            if was_archived:
                created += 1
            else:
                updated += 1

        db.session.commit()
        committed = True
    except Exception:
        db.session.rollback()
        raise
    finally:
        # A DB failure should not consume the preview token. Consume only after a
        # successful commit so the operator can correct the cause and retry.
        if committed:
            root = _stage_root()
            for suffix in (".xlsx", ".json"):
                candidate = root / f"{token}{suffix}"
                candidate.unlink(missing_ok=True)

    return {
        "created": created,
        "updated": updated,
        "month": month_key(selected_month),
    }


def prepare_next_month(user: User, source_month: date) -> dict[str, int | str]:
    target_month = add_months(source_month, 1)
    if target_month < current_month():
        raise PlanningError("Cannot prepare a month that is already in immutable history.")

    source = monthly_plans(user, source_month)
    target = _monthly_plan_slots(user, target_month)
    created = skipped = 0
    for committee in visible_committees(user):
        if not can_manage_action_plan(user, committee):
            continue
        if committee.id in target:
            skipped += 1
            continue
        previous = source.get(committee.id)
        plan = ActionPlan(
            committee_id=committee.id,
            title=committee.name,
            description=previous.description if previous else None,
            plan_month=target_month,
            plan_type=previous.plan_type if previous else None,
            assigned_date=None,
            assigned_by_user_id=user.id,
            prepared_from_id=previous.id if previous else None,
            notes=previous.notes if previous else None,
        )
        db.session.add(plan)
        db.session.flush()
        record_audit(AuditAction.CREATE, plan, after=model_snapshot(plan), actor=user)
        created += 1
    db.session.commit()
    return {
        "created": created,
        "skipped": skipped,
        "source_month": month_key(source_month),
        "target_month": month_key(target_month),
    }
