from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func

from ..extensions import db
from ..models import (
    DA,
    PC,
    PM,
    ActionPlan,
    AttendanceEntry,
    AuditAction,
    AuditLog,
    Cluster,
    Committee,
    CommitteeMember,
    SpecialsEntry,
    User,
    Village,
)


class WorkbookImportError(ValueError):
    pass


ENTITY_ORDER = ("pms", "pcs", "das", "villages", "committees", "members", "plans")


@dataclass
class SheetProfile:
    entity: str
    title: str
    header_row: int
    headers: dict[int, str]


def normalize(value: Any) -> str:
    text = str(value or "").strip().casefold()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def clean_phone(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return "".join(ch for ch in text if ch.isdigit() or ch == "+") or None


ALIASES: dict[str, dict[str, set[str]]] = {
    "pms": {
        "pm_code": {"pm id", "pm code"},
        "full_name": {"pm", "pm name", "name of pm", "program manager", "program manager name", "name"},
        "email": {"email", "email id", "mail id"},
        "mobile": {"mobile", "mobile number", "phone", "phone number", "contact", "contact number"},
        "notes": {"notes", "remarks"},
    },
    "pcs": {
        "pc_code": {"pc id", "pc code"},
        "full_name": {"pc", "pc name", "name of pc", "project coordinator", "project coordinator name", "name"},
        "cluster": {"cluster", "cluster name", "location", "centre", "center"},
        "email": {"email", "email id", "mail id"},
        "mobile": {"mobile", "mobile number", "phone", "phone number", "contact", "contact number"},
        "notes": {"notes", "remarks"},
    },
    "das": {
        "da_code": {"da id", "da code"},
        "full_name": {"da", "da name", "name of da", "development agent", "development agent name", "name"},
        "pc_code": {"pc id", "pc code"},
        "pc_name": {"pc", "pc name", "name of pc", "project coordinator", "project coordinator name"},
        "cluster": {"cluster", "cluster name", "location", "centre", "center"},
        "email": {"email", "email id", "mail id"},
        "mobile": {"mobile", "mobile number", "phone", "phone number", "contact", "contact number"},
        "notes": {"notes", "remarks"},
    },
    "villages": {
        "name": {"village", "village name", "name of village", "name"},
        "code": {"village code", "code", "village id"},
        "gp_name": {"gp", "gp name", "gram panchayat", "gram panchayat name"},
        "district": {"district", "district name"},
        "mandal": {"mandal", "mandal name", "block", "block name"},
        "latitude": {"latitude", "lat"},
        "longitude": {"longitude", "long", "lng", "lon"},
        "coordinates": {"coordinates", "gps", "gps coordinates", "geo coordinates"},
        "da_code": {"da id", "da code"},
        "da_name": {"da", "da name", "name of da", "development agent", "development agent name"},
        "pc_code": {"pc id", "pc code"},
        "pc_name": {"pc", "pc name", "project coordinator"},
        "cluster": {"cluster", "cluster name"},
        "notes": {"notes", "remarks"},
    },
    "committees": {
        "committee_code": {"committee id", "committee code"},
        "village_code": {"village id", "village code"},
        "name": {"committee", "committee name", "name of committee", "group", "group name", "name"},
        "committee_type": {"committee type", "type", "category", "sector"},
        "village_name": {"village", "village name", "name of village"},
        "da_name": {"da", "da name", "development agent"},
        "cluster": {"cluster", "cluster name"},
        "notes": {"notes", "remarks"},
    },
    "members": {
        "member_code": {"member id", "member code"},
        "committee_code": {"committee id", "committee code"},
        "village_code": {"village id", "village code"},
        "full_name": {"member", "member name", "name of member", "committee member", "committee member name", "name"},
        "committee_name": {"committee", "committee name", "name of committee", "group", "group name"},
        "village_name": {"village", "village name", "name of village"},
        "gender": {"gender", "sex"},
        "designation": {"designation", "position", "role", "committee role"},
        "designation_raw": {"designation raw", "raw designation"},
        "mobile": {"mobile", "mobile number", "phone", "phone number", "contact", "contact number"},
        "mobile_valid": {"mobile valid", "phone valid", "mobile validity"},
        "notes": {"notes", "remarks"},
    },
    "plans": {
        "plan_code": {"plan id", "plan code"},
        "village_code": {"village id", "village code"},
        "committee_code": {"committee id", "committee code"},
        "da_code": {"da id", "da code"},
        "plan_type": {"type", "plan type", "action plan type"},
        "frequency": {"frequency", "plan frequency"},
        "title": {"action plan", "action plan title", "plan", "plan title", "activity", "activity name", "title", "name"},
        "description": {"description", "details", "action", "planned action"},
        "assigned_date": {"assigned date", "assignment date", "target date", "planned date", "date", "visit date"},
        "committee_name": {"committee", "committee name", "name of committee", "group", "group name"},
        "village_name": {"village", "village name", "name of village"},
        "notes": {"notes", "remarks"},
    },
}


REQUIRED = {
    "pms": {"full_name"},
    "pcs": {"full_name", "cluster"},
    "das": {"full_name"},
    "villages": {"name"},
    "committees": {"name", "village_name"},
    "members": {"full_name", "committee_name", "village_name"},
    "plans": {"committee_name", "village_name"},
}


TITLE_HINTS = {
    "pms": ("pm", "program manager", "programme manager"),
    "pcs": ("pc", "project coordinator"),
    "das": ("da", "development agent"),
    "villages": ("village", "villages"),
    "committees": ("committee", "committees", "groups"),
    "members": ("member", "members"),
    "plans": ("action plan", "action plans", "plan", "plans"),
}


def _header_key(entity: str, value: Any) -> str | None:
    key = normalize(value)
    if not key:
        return None
    for canonical, aliases in ALIASES[entity].items():
        if key in aliases:
            return canonical
    # Tolerate numbered or decorated headings such as "Village Name *".
    for canonical, aliases in ALIASES[entity].items():
        if any(len(alias) >= 4 and (key.startswith(alias + " ") or key.endswith(" " + alias)) for alias in aliases):
            return canonical
    return None


def _title_score(entity: str, title: str) -> int:
    normalized = normalize(title)
    score = 0
    for hint in TITLE_HINTS[entity]:
        h = normalize(hint)
        if normalized == h:
            score += 10
        elif h in normalized:
            score += 4
    # Resolve committee-members before the broader committee hint.
    if entity == "committees" and "member" in normalized:
        score -= 8
    if entity == "plans" and "action" not in normalized and normalized == "plan":
        score += 2
    return score


def _profile_for_entity(ws: Worksheet, entity: str) -> SheetProfile | None:
    best: tuple[int, int, dict[int, str]] | None = None
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), start=1):
        headers: dict[int, str] = {}
        for index, value in enumerate(row):
            canonical = _header_key(entity, value)
            if canonical and canonical not in headers.values():
                headers[index] = canonical
        required_hits = len(REQUIRED[entity].intersection(headers.values()))
        score = required_hits * 12 + len(headers) * 2 + _title_score(entity, ws.title)
        if best is None or score > best[0]:
            best = (score, row_number, headers)
    if not best:
        return None
    score, header_row, headers = best
    if not REQUIRED[entity].issubset(headers.values()):
        return None
    return SheetProfile(entity, ws.title, header_row, headers)


def discover_sheets(workbook) -> dict[str, SheetProfile]:
    candidates: dict[str, list[tuple[int, SheetProfile]]] = defaultdict(list)
    for ws in workbook.worksheets:
        for entity in ENTITY_ORDER:
            profile = _profile_for_entity(ws, entity)
            if profile:
                score = _title_score(entity, ws.title) * 10 + len(profile.headers)
                candidates[entity].append((score, profile))

    selected: dict[str, SheetProfile] = {}
    used_titles: set[str] = set()
    # More specific entities first, then the dependency import order.
    choose_order = ("members", "plans", "committees", "villages", "das", "pcs", "pms")
    for entity in choose_order:
        options = sorted(candidates.get(entity, []), key=lambda item: item[0], reverse=True)
        unique = next((profile for _score, profile in options if profile.title not in used_titles), None)
        if unique:
            selected[entity] = unique
            used_titles.add(unique.title)

    missing = [entity for entity in ENTITY_ORDER if entity not in selected]
    # Action plans may be absent in a master workbook; all hierarchy sheets are mandatory.
    required_missing = [entity for entity in missing if entity != "plans"]
    if required_missing:
        readable = ", ".join(required_missing)
        raise WorkbookImportError(
            f"Could not identify required workbook sheets for: {readable}. "
            "Use the mapping option for non-standard headings."
        )
    return selected


def _iter_records(workbook, profile: SheetProfile) -> Iterable[tuple[int, dict[str, Any]]]:
    ws = workbook[profile.title]
    for row_number, row in enumerate(
        ws.iter_rows(min_row=profile.header_row + 1, values_only=True),
        start=profile.header_row + 1,
    ):
        record = {
            canonical: row[index] if index < len(row) else None
            for index, canonical in profile.headers.items()
        }
        if not any(value not in (None, "") for value in record.values()):
            continue
        yield row_number, record


def _cluster(value: Any, context: str) -> Cluster:
    normalized = normalize(value).replace(" ", "")
    if "csrb" in normalized:
        return Cluster.CSRB
    if "pdtc" in normalized:
        return Cluster.PDTC
    raise WorkbookImportError(f"{context}: cluster must be CSRB or PDTC, got {value!r}.")


def _excel_date(value: Any, context: str) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    raise WorkbookImportError(f"{context}: could not parse date {value!r}.")


def _coordinates(record: dict[str, Any]) -> tuple[float | None, float | None]:
    lat, lon = record.get("latitude"), record.get("longitude")
    if (lat in (None, "")) and (lon in (None, "")) and record.get("coordinates"):
        pieces = re.findall(r"-?\d+(?:\.\d+)?", str(record["coordinates"]))
        if len(pieces) >= 2:
            lat, lon = pieces[:2]
    try:
        latitude = float(lat) if lat not in (None, "") else None
        longitude = float(lon) if lon not in (None, "") else None
    except (TypeError, ValueError) as exc:
        raise WorkbookImportError(f"Invalid village coordinates: {lat!r}, {lon!r}.") from exc
    if latitude is not None and not -90 <= latitude <= 90:
        raise WorkbookImportError(f"Latitude {latitude} is outside -90..90.")
    if longitude is not None and not -180 <= longitude <= 180:
        raise WorkbookImportError(f"Longitude {longitude} is outside -180..180.")
    return latitude, longitude


def _unique_by_name(items: Iterable[Any], name: Any, context: str):
    wanted = normalize(name)
    matches = [item for item in items if normalize(getattr(item, "full_name", getattr(item, "name", ""))) == wanted]
    if len(matches) == 1:
        return matches[0]
    if not matches:
        raise WorkbookImportError(f"{context}: no matching record for {name!r}.")
    raise WorkbookImportError(f"{context}: {name!r} is ambiguous; include its parent context.")


def _validate_cluster(row_cluster: Any, inherited: Cluster, context: str) -> None:
    if row_cluster in (None, ""):
        return
    supplied = _cluster(row_cluster, context)
    if supplied != inherited:
        raise WorkbookImportError(
            f"{context}: workbook cluster {supplied.value} conflicts with inherited cluster {inherited.value}."
        )


def _load_mapping(path: str | Path | None) -> dict[str, Any]:
    if not path:
        return {}
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise WorkbookImportError(f"Could not read mapping file: {exc}") from exc


def _profiles_from_mapping(workbook, mapping: dict[str, Any]) -> dict[str, SheetProfile]:
    if not mapping:
        return discover_sheets(workbook)
    profiles: dict[str, SheetProfile] = {}
    for entity in ENTITY_ORDER:
        config = mapping.get(entity)
        if not config:
            if entity == "plans":
                continue
            raise WorkbookImportError(f"Mapping is missing {entity}.")
        title = config["sheet"]
        if title not in workbook.sheetnames:
            raise WorkbookImportError(f"Mapped sheet {title!r} was not found.")
        header_row = int(config.get("header_row", 1))
        ws = workbook[title]
        row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        explicit = {normalize(key): value for key, value in config.get("columns", {}).items()}
        headers: dict[int, str] = {}
        for index, value in enumerate(row):
            normalized = normalize(value)
            canonical = explicit.get(normalized) or _header_key(entity, value)
            if canonical:
                headers[index] = canonical
        if not REQUIRED[entity].issubset(headers.values()):
            raise WorkbookImportError(f"Mapped {entity} sheet is missing required columns.")
        profiles[entity] = SheetProfile(entity, title, header_row, headers)
    return profiles


def _pc_da_assignments(workbook) -> tuple[dict[str, str], dict[str, str]]:
    """Read the workbook's authoritative PC_DA_Map without formula caches.

    Returns two normalized lookup dictionaries:
    - DA_ID -> PC_ID
    - DA_Name -> PC_ID

    The sheet is optional for backwards compatibility. When present, conflicting
    duplicate mappings are rejected instead of silently choosing one.
    """
    sheet_title = next(
        (title for title in workbook.sheetnames if normalize(title) == "pc da map"),
        None,
    )
    if sheet_title is None:
        return {}, {}

    ws = workbook[sheet_title]
    header_row: int | None = None
    pc_index: int | None = None
    da_index: int | None = None
    da_name_index: int | None = None

    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True),
        start=1,
    ):
        normalized_headers = {
            normalize(value): index
            for index, value in enumerate(row)
            if normalize(value)
        }

        candidate_pc = normalized_headers.get("pc id")
        candidate_da = normalized_headers.get("da id")

        if candidate_pc is not None and candidate_da is not None:
            header_row = row_number
            pc_index = candidate_pc
            da_index = candidate_da
            da_name_index = normalized_headers.get("da name")
            break

    if header_row is None or pc_index is None or da_index is None:
        raise WorkbookImportError(
            f"{sheet_title}: expected PC_ID and DA_ID columns."
        )

    by_da_code: dict[str, str] = {}
    by_da_name: dict[str, str] = {}

    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        pc_code = clean_text(row[pc_index]) if pc_index < len(row) else None
        da_code = clean_text(row[da_index]) if da_index < len(row) else None
        da_name = (
            clean_text(row[da_name_index])
            if da_name_index is not None and da_name_index < len(row)
            else None
        )

        if not pc_code and not da_code and not da_name:
            continue

        if not pc_code or not da_code:
            raise WorkbookImportError(
                f"{sheet_title} row {row_number}: PC_ID and DA_ID are required."
            )

        da_key = normalize(da_code)
        existing_pc = by_da_code.get(da_key)
        if existing_pc and normalize(existing_pc) != normalize(pc_code):
            raise WorkbookImportError(
                f"{sheet_title} row {row_number}: DA_ID {da_code!r} is mapped "
                f"to more than one PC."
            )
        by_da_code[da_key] = pc_code

        if da_name:
            name_key = normalize(da_name)
            existing_name_pc = by_da_name.get(name_key)
            if existing_name_pc and normalize(existing_name_pc) != normalize(pc_code):
                raise WorkbookImportError(
                    f"{sheet_title} row {row_number}: DA_Name {da_name!r} is "
                    f"mapped to more than one PC."
                )
            by_da_name[name_key] = pc_code

    return by_da_code, by_da_name


def database_counts() -> dict[str, int]:
    models = {
        "pms": PM,
        "pcs": PC,
        "das": DA,
        "villages": Village,
        "committees": Committee,
        "members": CommitteeMember,
        "plans": ActionPlan,
    }
    return {
        key: db.session.scalar(db.select(func.count(model.id)).where(model.is_deleted.is_(False))) or 0
        for key, model in models.items()
    }


def _prepare_replace() -> None:
    if (db.session.scalar(db.select(func.count(AttendanceEntry.id))) or 0) > 0:
        raise WorkbookImportError("Replacement is blocked because attendance entries exist.")
    if (db.session.scalar(db.select(func.count(SpecialsEntry.id))) or 0) > 0:
        raise WorkbookImportError("Replacement is blocked because specials entries exist.")
    if (
        db.session.scalar(
            db.select(func.count(User.id)).where(
                (User.pm_id.is_not(None)) | (User.pc_id.is_not(None)) | (User.da_id.is_not(None))
            )
        )
        or 0
    ) > 0:
        raise WorkbookImportError("Replacement is blocked while role-profile users exist.")

    for model in (ActionPlan, CommitteeMember, Committee, Village, DA, PC, PM):
        db.session.execute(db.delete(model))


def import_workbook(
    path: str | Path,
    *,
    replace: bool = False,
    mapping_path: str | Path | None = None,
    actor_user_id: int | None = None,
) -> dict[str, Any]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise WorkbookImportError(f"Workbook not found: {workbook_path}")
    if workbook_path.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise WorkbookImportError("Master data must be an .xlsx or .xlsm workbook.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    mapping = _load_mapping(mapping_path)
    profiles = _profiles_from_mapping(workbook, mapping)
    existing = database_counts()
    if any(existing.values()) and not replace:
        raise WorkbookImportError(
            "Master data already exists. Re-run with --replace only after taking a backup "
            "and confirming there are no field entries or profile users."
        )

    summary: dict[str, Any] = {"source": workbook_path.name, "sheets": {}, "counts": Counter()}
    if replace:
        _prepare_replace()
        db.session.flush()

    try:
        pms: list[PM] = []
        for _row_no, row in _iter_records(workbook, profiles["pms"]):
            name = clean_text(row.get("full_name"))
            if not name:
                continue
            pm = PM(
                full_name=name,
                email=clean_text(row.get("email")).casefold() if clean_text(row.get("email")) else None,
                mobile=clean_phone(row.get("mobile")),
                notes=clean_text(row.get("notes")),
            )
            db.session.add(pm)
            pms.append(pm)
        db.session.flush()
        summary["counts"]["pms"] = len(pms)

        pcs: list[PC] = []
        pc_by_code: dict[str, PC] = {}
        for row_no, row in _iter_records(workbook, profiles["pcs"]):
            name = clean_text(row.get("full_name"))
            if not name:
                continue
            context = f"{profiles['pcs'].title} row {row_no}"
            pc = PC(
                full_name=name,
                cluster=_cluster(row.get("cluster"), context),
                email=clean_text(row.get("email")).casefold() if clean_text(row.get("email")) else None,
                mobile=clean_phone(row.get("mobile")),
                notes=clean_text(row.get("notes")),
            )
            db.session.add(pc)
            pcs.append(pc)
            code = clean_text(row.get("pc_code"))
            if code:
                key = normalize(code)
                if key in pc_by_code:
                    raise WorkbookImportError(f"{context}: duplicate PC_ID {code!r}.")
                pc_by_code[key] = pc
        db.session.flush()
        summary["counts"]["pcs"] = len(pcs)

        pc_code_by_da_code, pc_code_by_da_name = _pc_da_assignments(workbook)

        das: list[DA] = []
        da_by_code: dict[str, DA] = {}
        for row_no, row in _iter_records(workbook, profiles["das"]):
            name = clean_text(row.get("full_name"))
            if not name:
                continue

            context = f"{profiles['das'].title} row {row_no}"
            da_code = clean_text(row.get("da_code"))
            sheet_pc_code = clean_text(row.get("pc_code"))

            mapped_pc_code = None
            if da_code:
                mapped_pc_code = pc_code_by_da_code.get(normalize(da_code))
            if mapped_pc_code is None:
                mapped_pc_code = pc_code_by_da_name.get(normalize(name))

            if (
                sheet_pc_code
                and mapped_pc_code
                and normalize(sheet_pc_code) != normalize(mapped_pc_code)
            ):
                raise WorkbookImportError(
                    f"{context}: PC_ID {sheet_pc_code!r} conflicts with "
                    f"PC_DA_Map assignment {mapped_pc_code!r}."
                )

            # PC_DA_Map is authoritative when present. The DA-sheet PC_ID/PC_Name
            # remains a backwards-compatible fallback for older workbooks.
            pc_code = mapped_pc_code or sheet_pc_code
            pc = pc_by_code.get(normalize(pc_code)) if pc_code else None

            if pc is None and row.get("pc_name"):
                pc = _unique_by_name(pcs, row.get("pc_name"), context)

            if pc is None:
                raise WorkbookImportError(
                    f"{context}: no matching PC for "
                    f"{pc_code or row.get('pc_name')!r}. "
                    "Provide PC_DA_Map or a populated DA PC_ID/PC_Name."
                )

            _validate_cluster(row.get("cluster"), pc.cluster, context)

            da = DA(
                full_name=name,
                pc_id=pc.id,
                email=clean_text(row.get("email")).casefold()
                if clean_text(row.get("email"))
                else None,
                mobile=clean_phone(row.get("mobile")),
                notes=clean_text(row.get("notes")),
            )
            db.session.add(da)
            das.append(da)

            if da_code:
                key = normalize(da_code)
                if key in da_by_code:
                    raise WorkbookImportError(
                        f"{context}: duplicate DA_ID {da_code!r}."
                    )
                da_by_code[key] = da
        db.session.flush()
        summary["counts"]["das"] = len(das)

        villages: list[Village] = []
        village_index: dict[str, list[Village]] = defaultdict(list)
        for row_no, row in _iter_records(workbook, profiles["villages"]):
            name = clean_text(row.get("name"))
            if not name:
                continue
            context = f"{profiles['villages'].title} row {row_no}"
            da_code = clean_text(row.get("da_code"))
            da = da_by_code.get(normalize(da_code)) if da_code else None
            if da is None and row.get("da_name"):
                da = _unique_by_name(das, row.get("da_name"), context)
            if da is None:
                raise WorkbookImportError(f"{context}: no matching DA for {da_code or row.get('da_name')!r}.")
            _validate_cluster(row.get("cluster"), da.pc.cluster, context)
            if row.get("pc_name"):
                pc = _unique_by_name(pcs, row.get("pc_name"), context)
                if pc.id != da.pc_id:
                    raise WorkbookImportError(f"{context}: PC does not match the selected DA.")
            latitude, longitude = _coordinates(row)
            village = Village(
                name=name,
                code=clean_text(row.get("code")),
                gp_name=clean_text(row.get("gp_name")),
                district=clean_text(row.get("district")),
                mandal=clean_text(row.get("mandal")),
                latitude=latitude,
                longitude=longitude,
                da_id=da.id,
                notes=clean_text(row.get("notes")),
            )
            db.session.add(village)
            villages.append(village)
            village_index[normalize(name)].append(village)
        db.session.flush()
        summary["counts"]["villages"] = len(villages)

        committees: list[Committee] = []
        committee_index: dict[tuple[str, str], list[Committee]] = defaultdict(list)
        for row_no, row in _iter_records(workbook, profiles["committees"]):
            name = clean_text(row.get("name"))
            if not name:
                continue
            context = f"{profiles['committees'].title} row {row_no}"
            village = _unique_by_name(villages, row.get("village_name"), context)
            _validate_cluster(row.get("cluster"), village.da.pc.cluster, context)
            committee = Committee(
                name=name,
                committee_type=clean_text(row.get("committee_type")),
                village_id=village.id,
                notes=clean_text(row.get("notes")),
            )
            db.session.add(committee)
            committees.append(committee)
            committee_index[(normalize(village.name), normalize(name))].append(committee)
        db.session.flush()
        summary["counts"]["committees"] = len(committees)

        members: list[CommitteeMember] = []
        for row_no, row in _iter_records(workbook, profiles["members"]):
            name = clean_text(row.get("full_name"))
            if not name:
                continue
            context = f"{profiles['members'].title} row {row_no}"
            key = (normalize(row.get("village_name")), normalize(row.get("committee_name")))
            matches = committee_index.get(key, [])
            if len(matches) != 1:
                raise WorkbookImportError(f"{context}: committee/village reference is missing or ambiguous.")
            designation = clean_text(row.get("designation"))
            raw_designation = clean_text(row.get("designation_raw"))
            notes = clean_text(row.get("notes"))
            if raw_designation and raw_designation != designation:
                raw_note = f"Source designation: {raw_designation}"
                notes = f"{notes}\n{raw_note}" if notes else raw_note
            mobile_valid = normalize(row.get("mobile_valid"))
            mobile = clean_phone(row.get("mobile"))
            if mobile_valid and mobile_valid not in {"valid", "yes", "true", "1"}:
                mobile = None
            member = CommitteeMember(
                committee_id=matches[0].id,
                full_name=name,
                gender=clean_text(row.get("gender")),
                designation=designation,
                mobile=mobile,
                notes=notes,
            )
            db.session.add(member)
            members.append(member)
        db.session.flush()
        summary["counts"]["members"] = len(members)

        plans: list[ActionPlan] = []
        if "plans" in profiles:
            for row_no, row in _iter_records(workbook, profiles["plans"]):
                village_name = clean_text(row.get("village_name"))
                committee_name = clean_text(row.get("committee_name"))
                if not village_name and not committee_name:
                    continue
                context = f"{profiles['plans'].title} row {row_no}"
                key = (normalize(village_name), normalize(committee_name))
                matches = committee_index.get(key, [])
                if len(matches) != 1:
                    raise WorkbookImportError(f"{context}: committee/village reference is missing or ambiguous.")
                plan_code = clean_text(row.get("plan_code"))
                title = clean_text(row.get("title")) or committee_name or plan_code or f"Action Plan {row_no}"
                plan = ActionPlan(
                    committee_id=matches[0].id,
                    title=title,
                    description=clean_text(row.get("description")),
                    # Workbook rows are templates. Monthly executable occurrences are
                    # created later by PC planning; do not silently turn source dates
                    # into an assigned monthly schedule.
                    plan_month=None,
                    plan_type=None,
                    assigned_date=None,
                    notes=clean_text(row.get("notes")),
                )
                db.session.add(plan)
                plans.append(plan)
            db.session.flush()
        summary["counts"]["plans"] = len(plans)

        summary["sheets"] = {
            entity: {
                "title": profile.title,
                "header_row": profile.header_row,
                "columns": sorted(profile.headers.values()),
            }
            for entity, profile in profiles.items()
        }
        summary["counts"] = dict(summary["counts"])
        db.session.add(
            AuditLog(
                actor_user_id=actor_user_id,
                action=AuditAction.IMPORT,
                entity_type="Workbook",
                entity_id=None,
                after_json=summary,
            )
        )
        db.session.commit()
        return summary
    except Exception:
        db.session.rollback()
        raise
    finally:
        workbook.close()
