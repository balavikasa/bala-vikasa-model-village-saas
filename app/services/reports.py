from __future__ import annotations

from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import selectinload

from ..extensions import db
from ..models import (
    ActionPlan,
    ActionPlanType,
    AttendanceEntry,
    Committee,
    CommitteeMember,
    DA,
    SpecialsEntry,
    User,
    Village,
)
from ..scoping import inherited_cluster, scoped_select
from .monthly_plans import action_plan_status, month_key, month_label


def _month_bounds(month: date) -> tuple[date, date]:
    if month.month == 12:
        nxt = date(month.year + 1, 1, 1)
    else:
        nxt = date(month.year, month.month + 1, 1)
    return month, nxt


def report_plans(
    user: User,
    month: date,
    *,
    plan_type: str | None = None,
    status: str | None = None,
) -> list[ActionPlan]:
    stmt = (
        scoped_select(ActionPlan, user)
        .where(
            ActionPlan.plan_month == month,
            ActionPlan.plan_type.is_not(None),
            ActionPlan.assigned_date.is_not(None),
        )
        .options(
            selectinload(ActionPlan.committee)
            .selectinload(Committee.village)
            .selectinload(Village.da)
            .selectinload(DA.pc),
            selectinload(ActionPlan.attendance_entry)
            .selectinload(AttendanceEntry.visited_members),
            selectinload(ActionPlan.specials_entry),
        )
        .order_by(ActionPlan.assigned_date.desc(), ActionPlan.id.desc())
    )
    if plan_type:
        stmt = stmt.where(ActionPlan.plan_type == ActionPlanType(plan_type))
    plans = list(db.session.scalars(stmt).unique())
    if status:
        plans = [plan for plan in plans if action_plan_status(plan) == status]
    return plans


def report_row(plan: ActionPlan) -> dict[str, Any]:
    village = plan.committee.village
    da = village.da
    entry = plan.attendance_entry if plan.plan_type == ActionPlanType.ATTENDANCE else plan.specials_entry
    activity_date = None
    if isinstance(entry, AttendanceEntry):
        activity_date = entry.visit_date
    elif isinstance(entry, SpecialsEntry):
        activity_date = entry.event_date
    return {
        "plan_id": plan.id,
        "date": (activity_date or plan.assigned_date).isoformat() if (activity_date or plan.assigned_date) else None,
        "assigned_date": plan.assigned_date.isoformat() if plan.assigned_date else None,
        "da": da.full_name,
        "village": village.name,
        "committee": plan.committee.name,
        "type": plan.plan_type.value if plan.plan_type else "Draft",
        "status": action_plan_status(plan),
        "has_entry": bool(entry and not entry.is_deleted),
        "view_url": f"/reports/plan/{plan.id}",
    }


def report_rows(user: User, month: date, *, plan_type: str | None = None, status: str | None = None):
    return [report_row(plan) for plan in report_plans(user, month, plan_type=plan_type, status=status)]


def report_detail(plan: ActionPlan) -> dict[str, Any]:
    committee = plan.committee
    village = committee.village
    da = village.da
    pc = da.pc
    counts = dict(
        db.session.execute(
            db.select(CommitteeMember.gender, func.count(CommitteeMember.id))
            .where(
                CommitteeMember.committee_id == committee.id,
                CommitteeMember.is_enabled.is_(True),
                CommitteeMember.is_deleted.is_(False),
            )
            .group_by(CommitteeMember.gender)
        ).all()
    )
    detail: dict[str, Any] = {
        "plan_id": plan.id,
        "month": month_key(plan.plan_month) if plan.plan_month else None,
        "month_label": month_label(plan.plan_month) if plan.plan_month else None,
        "type": plan.plan_type.value if plan.plan_type else "Draft",
        "status": action_plan_status(plan),
        "assigned_date": plan.assigned_date.isoformat() if plan.assigned_date else None,
        "notes": plan.notes,
        "cluster": inherited_cluster(plan),
        "pc": pc.full_name,
        "da": da.full_name,
        "village": village.name,
        "mandal": village.mandal,
        "district": village.district,
        "committee": committee.name,
        "latitude": village.latitude,
        "longitude": village.longitude,
        "master": {
            "male": counts.get("Male", 0),
            "female": counts.get("Female", 0),
            "unknown": sum(value for key, value in counts.items() if key not in {"Male", "Female"}),
            "total": sum(counts.values()),
        },
        "entry": None,
    }
    if plan.attendance_entry and not plan.attendance_entry.is_deleted:
        entry = plan.attendance_entry
        grouped: dict[str, list[str]] = defaultdict(list)
        for selected in entry.visited_members:
            if selected.is_enabled and not selected.is_deleted:
                grouped[selected.designation_snapshot].append(selected.member_name_snapshot)
        detail["latitude"] = entry.latitude if entry.latitude is not None else detail["latitude"]
        detail["longitude"] = entry.longitude if entry.longitude is not None else detail["longitude"]
        detail["entry"] = {
            "id": entry.id,
            "date": entry.visit_date.isoformat(),
            "male": entry.male_count,
            "female": entry.female_count,
            "total": entry.total_count,
            "new_members": entry.new_members_count,
            "visit_members": dict(grouped),
            "reason": entry.reason,
            "remarks": entry.remarks,
            "photo_url": f"/api/v1/photos/attendance/{entry.id}" if entry.photo_path else None,
            "gps_source": entry.geolocation_source,
            "submitted_at": entry.submitted_at.isoformat(),
            "submitted_by": entry.submitted_by.display_name,
        }
    elif plan.specials_entry and not plan.specials_entry.is_deleted:
        entry = plan.specials_entry
        detail["latitude"] = entry.latitude if entry.latitude is not None else detail["latitude"]
        detail["longitude"] = entry.longitude if entry.longitude is not None else detail["longitude"]
        detail["entry"] = {
            "id": entry.id,
            "date": entry.event_date.isoformat(),
            "title": entry.title or committee.name,
            "participants": entry.participant_count,
            "scope": entry.scope.value,
            "reason": entry.reason,
            "notes": entry.notes,
            "photo_url": f"/api/v1/photos/specials/{entry.id}" if entry.photo_path else None,
            "gps_source": entry.geolocation_source,
            "submitted_at": entry.submitted_at.isoformat(),
            "submitted_by": entry.submitted_by.display_name,
        }
    return detail


def report_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    statuses = {"Early": 0, "On-time": 0, "Postponed": 0, "Failure": 0, "Scheduled": 0, "Due today": 0}
    for row in rows:
        if row["status"] in statuses:
            statuses[row["status"]] += 1
    return {"total": len(rows), **statuses}


def export_report_workbook(user: User, month: date, *, plan_type: str | None = None, status: str | None = None) -> Workbook:
    plans = report_plans(user, month, plan_type=plan_type, status=status)
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    headers = [
        "Month", "Date", "Assigned Date", "Cluster", "PC", "DA", "Village", "Committee",
        "Type", "Status", "Male", "Female", "Total", "New Members",
        "Committee Member Name", "Visit Designation", "Participants", "Scope", "Reason",
        "Notes / Remarks", "Latitude", "Longitude", "Photo URL",
    ]
    ws.append(headers)
    for plan in plans:
        detail = report_detail(plan)
        entry = detail["entry"] or {}
        visit_members = entry.get("visit_members") or {}
        flattened = [
            (designation, name)
            for designation, names in visit_members.items()
            for name in names
        ] or [("", "")]
        for index, (designation, member_name) in enumerate(flattened):
            ws.append(
                [
                    detail["month"],
                    entry.get("date") if index == 0 else entry.get("date"),
                    detail["assigned_date"],
                    detail["cluster"],
                    detail["pc"],
                    detail["da"],
                    detail["village"],
                    detail["committee"],
                    detail["type"],
                    detail["status"],
                    entry.get("male"),
                    entry.get("female"),
                    entry.get("total"),
                    entry.get("new_members"),
                    member_name,
                    designation,
                    entry.get("participants"),
                    entry.get("scope"),
                    entry.get("reason"),
                    entry.get("remarks") or entry.get("notes") or detail["notes"],
                    detail["latitude"],
                    detail["longitude"],
                    entry.get("photo_url"),
                ]
            )

    navy = "1F4E5F"
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    widths = [12, 14, 16, 12, 22, 22, 24, 28, 14, 14, 10, 10, 10, 14, 26, 18, 14, 14, 28, 34, 14, 14, 28]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    return wb
