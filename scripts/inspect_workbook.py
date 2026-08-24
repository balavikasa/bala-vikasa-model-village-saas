#!/usr/bin/env python3
"""Profile the supplied workbook without changing the database."""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def profile(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, object] = {}
    for sheet in workbook.worksheets:
        header_row = 1
        headers: list[str] = []
        for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(20, sheet.max_row), values_only=True), 1):
            normalized = [normalize(cell) for cell in row]
            populated = [cell for cell in normalized if cell]
            if len(populated) >= 2 and len(set(populated)) == len(populated):
                header_row = index
                headers = normalized
                break
        records = 0
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if any(value not in (None, "") for value in row):
                records += 1
        result[sheet.title] = {
            "header_row": header_row,
            "headers": headers,
            "records": records,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
        }
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    data = profile(args.workbook)
    payload = json.dumps(data, indent=2, ensure_ascii=False)
    if args.output:
        args.output.write_text(payload + "\n", encoding="utf-8")
    else:
        print(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
